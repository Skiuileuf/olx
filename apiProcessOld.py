from datetime import datetime, timezone
import json
from openpyxl import Workbook
from openpyxl.utils import datetime as openpyxl_datetime
from openpyxl.styles import NamedStyle, numbers
from typing import Callable, Any, Optional
import glob

def iso_to_excel_date(iso_string: str) -> float:
    """
    Converts an ISO formatted date string to an Excel date.
    Args:
        iso_string (str): The ISO formatted date string.
    Returns:
        float: The Excel date representation.
    """
    dt = datetime.fromisoformat(iso_string)
    
    # Convert to UTC
    dt_utc = dt.astimezone(timezone.utc)
    
    # Remove timezone information to make it naive
    dt_naive = dt_utc.replace(tzinfo=None)
    
    # Convert to Excel date
    return openpyxl_datetime.to_excel(dt_naive)


def get_param_value(params, key: str) -> dict:
    for param in params:
        if param['key'] == key:
            return param['value']
    return None  # Return None if the key is not found


wb = Workbook()
ws = wb.active

# Create a new named style for the date format
date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM:SS')

# Add the new style to the workbook
wb.add_named_style(date_style)

# Otopeni coordinates
otopeni_lat = 44.55048
otopeni_lon = 26.09095

current_row: int = 1

header_row = ["id", "url", "title", "last_refresh_time", "created_time", "description", "price", "currency", "price_equiv", "distance", "city", "region", "days_since_last_refresh", "days_since_created_time", "highlighted"]
current_row += 1

files = glob.glob("api/*.json")

params_values = []
for filePath in files:
    f = open(filePath, "rb")
    data = f.read()
    python_obj = json.loads(data)
    for item in python_obj["data"]:
        params = item["params"]
        for param in params:
            if param["key"] not in params_values:
                params_values.append(param["key"])

print(params_values)

header_row += params_values

ws.append(header_row)

for filePath in files:
    print(filePath)
    f = open(filePath, "rb")
    data = f.read()

    python_obj = json.loads(data)

    for item in python_obj["data"]:
        row = []
        row.append(item["id"])
        row.append(item["url"])
        row.append(item["title"])
        row.append(iso_to_excel_date(item["last_refresh_time"]))
        row.append(iso_to_excel_date(item["created_time"]))
        row.append(item["description"])
        
        # state = get_param_value(item["params"], "state")
        # diagonala = get_param_value(item["params"], "diagonala")
        # producator_procesor = get_param_value(item["params"], "producator_procesor")
        # capacitate_memorie_ram = get_param_value(item["params"], "capacitate_memorie_ram")
        price = get_param_value(item["params"], "price")

        # row.append(state["label"] if state else None)
        # row.append(diagonala["label"] if diagonala else None)
        # row.append(producator_procesor["label"] if producator_procesor else None)
        # row.append(capacitate_memorie_ram["label"] if capacitate_memorie_ram else None)
        row.append(price["value"] if price else None)
        row.append(price["currency"] if price else None)


        #priceequiv
        price_equiv = None
        if price:
            if price["currency"] != "RON":
                price_equiv = price["converted_value"]
            else:
                price_equiv = price["value"]

        row.append(price_equiv)
        # row.append(price["negotiable"] if price else None)

        lat = item["map"]["lat"]
        lon = item["map"]["lon"]

        # row.append(lat)
        # row.append(lon)

        #calculate synthetic indicator representing straight line distance from Otopeni to the location
        distance = ((lat - otopeni_lat)**2 + (lon - otopeni_lon)**2)**0.5
        row.append(distance)

        row.append(item["location"]["city"]["name"])
        row.append(item["location"]["region"]["name"])

        row.append(f"=DATEDIF(D{current_row}, TODAY() , \"D\")")
        row.append(f"=DATEDIF(E{current_row}, TODAY() , \"D\")")

        row.append(item["promotion"]["highlighted"])

        for i in range(len(params_values)):
            param_value = get_param_value(item["params"], params_values[i])
            if param_value:
                row.append(param_value["label"])
            else:
                row.append(None)
    
        ws.append(row)

        ws.cell(row=current_row, column=2).style = "Hyperlink"
        ws.cell(row=current_row, column=2).hyperlink = item["url"]

        ws.cell(row=current_row, column=4).style = date_style
        ws.cell(row=current_row, column=5).style = date_style

        
        current_row += 1

wb.save("olx-0-50-macbook+pro.xlsx")